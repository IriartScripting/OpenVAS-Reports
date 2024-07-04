import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
import os
from datetime import datetime

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

class ReportManager:
    def __init__(self, filters_file, smtp_server, smtp_port, smtp_user, smtp_password):
        self.filters_file = "./filters/" + filters_file
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.smtp_user = smtp_user
        self.smtp_password = smtp_password

    def read_filter_ids(self):
            """
            Lee un archivo línea por línea para obtener los IDs de los filtros y sus nombres.
            """
            filters = []
            try:
                with open(self.filters_file, 'r') as file:
                    for line in file:
                        line = line.strip()
                        if line:  # Evitar líneas vacías
                            nombre_filtro, filter_id = line.split(',')
                            filters.append((nombre_filtro.strip(), filter_id.strip()))
            except FileNotFoundError:
                print(f"El archivo {self.filters_file} no se encontró.")
            except Exception as e:
                print(f"Ocurrió un error leyendo el archivo {self.filters_file}: {e}")
            
            return filters


    def csvs_to_excel(self, csv_folder, excel_filename, delimiter=','):
        """
        Combina múltiples archivos CSV en un solo archivo Excel, donde cada CSV
        se convierte en una hoja separada.

        Args:
        - csv_folder (str): Directorio donde se encuentran los archivos CSV.
        - excel_filename (str): Nombre del archivo Excel de salida.
        - delimiter (str): Delimitador usado en los archivos CSV (por defecto es coma).
        """
        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            for csv_filename in os.listdir(csv_folder):
                if csv_filename.endswith('.csv'):
                    csv_path = os.path.join(csv_folder, csv_filename)
                    df = pd.read_csv(csv_path, delimiter=delimiter)
                    sheet_name = os.path.splitext(csv_filename)[0]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"CSV files successfully combined into {excel_filename}")
            
                    # Leer el archivo Excel combinado
        df_combined = pd.read_excel(excel_filename, sheet_name=None)

        # Procesar cada hoja del archivo Excel
        for sheet_name, df in df_combined.items():
            # Conservar solo las columnas 'IP' y 'Hostname'
            if 'IP' in df.columns and 'Hostname' in df.columns:
                df = df[['IP', 'Hostname']]

                # Eliminar filas duplicadas basadas en 'IP' y 'Hostname'
                df = df.drop_duplicates(subset=['IP', 'Hostname'])
                
                # Guardar el DataFrame filtrado de vuelta en el archivo Excel
                with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a') as writer:
                    if sheet_name in writer.book.sheetnames:
                        # Si la hoja ya existe, eliminarla antes de escribir la nueva versión
                        writer.book.remove(writer.book[sheet_name])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

    def add_summary_chart(self, excel_filename):
        # Cargar el archivo Excel existente
        wb = load_workbook(excel_filename)
        
        # Verificar si la hoja 'Summary' existe, si no, crearla
        if 'Summary' in wb.sheetnames:
            wb.remove(wb['Summary'])
        summary_sheet = wb.create_sheet(title='Summary', index=0)
        
        # Definir una función para obtener el conteo de IPs y Hostnames para cada hoja
        def get_ip_hostname_counts(sheet):
            df = pd.DataFrame(sheet.values)
            df.columns = df.iloc[0]
            df = df[1:]
            ip_count = df['IP'].nunique()
            hostname_count = df['Hostname'].nunique()
            return ip_count, hostname_count

        # Crear el DataFrame para el resumen
        summary_data = {
            'Sheet Name': [],
            'IP Count': [],
            'Hostname Count': []
        }

        for sheet_name in wb.sheetnames:
            if sheet_name == 'Summary':
                continue
            sheet = wb[sheet_name]
            ip_count, hostname_count = get_ip_hostname_counts(sheet)
            summary_data['Sheet Name'].append(sheet_name)
            summary_data['IP Count'].append(ip_count)
            summary_data['Hostname Count'].append(hostname_count)

        summary_df = pd.DataFrame(summary_data)

        # Agregar los datos del resumen a la hoja 'Summary'
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = summary_sheet.cell(row=r_idx, column=c_idx, value=value)
                # Aplicar estilos a la cabecera
                if r_idx == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=Side(border_style="thin", color='000000'),
                                        right=Side(border_style="thin", color='000000'),
                                        top=Side(border_style="thin", color='000000'),
                                        bottom=Side(border_style="thin", color='000000'))
                else:
                    cell.border = Border(left=Side(border_style="thin", color='000000'),
                                        right=Side(border_style="thin", color='000000'),
                                        top=Side(border_style="thin", color='000000'),
                                        bottom=Side(border_style="thin", color='000000'))

        # Crear el gráfico
        chart = BarChart()
        chart.title = "IP and Hostname Counts by Sheet"
        chart.style = 10
        chart.x_axis.title = 'Sheet Name'
        chart.y_axis.title = 'Count'
        chart.y_axis.majorGridlines = None

        # Agregar datos al gráfico
        data = Reference(summary_sheet, min_col=2, min_row=1, max_col=3, max_row=len(summary_df) + 1)
        categories = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(summary_df) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        # Colocar el gráfico en la hoja de resumen
        summary_sheet.add_chart(chart, "E5")

        # Obtener la fecha actual y agregarla a la hoja
        today = datetime.today().strftime('%Y-%m-%d')
        summary_sheet.cell(row=1, column=5, value=f"Report Date: {today}")
        summary_sheet.cell(row=1, column=5).font = Font(bold=True)

        # Guardar el archivo Excel con la fecha actual en el nombre
        new_excel_filename = f"{excel_filename}_{today}.xlsx"
        wb.save(new_excel_filename)
        wb.close()
        print(f"Summary chart added and saved as {new_excel_filename}")

    def send_email(self, to_address, subject, body, attachment_path):
        """
        Envía un correo electrónico con el contenido especificado y adjunta un archivo.
        
        Args:
        - to_address (str): Dirección de correo del destinatario.
        - subject (str): Asunto del correo electrónico.
        - body (str): Cuerpo del mensaje del correo electrónico.
        - attachment_path (str): Ruta del archivo que se va a adjuntar.
        """
        try:
            # Crear el mensaje
            msg = MIMEMultipart()
            msg['From'] = self.smtp_user
            msg['To'] = to_address
            msg['Subject'] = subject

            # Adjuntar el cuerpo del mensaje
            msg.attach(MIMEText(body, 'plain'))

            # Adjuntar el archivo
            with open(attachment_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {attachment_path.split("/")[-1]}'
                )
                msg.attach(part)

            # Configurar el servidor SMTP
            server = smtplib.SMTP(host=self.smtp_server, port=self.smtp_port)
            server.starttls()  # Usar TLS
            #server.login(self.smtp_user, self.smtp_password)

            # Enviar el correo
            server.send_message(msg)
            server.quit()
            print(f"Correo enviado a {to_address}")
        
        except Exception as e:
            print(f"Ocurrió un error enviando el correo: {e}")


